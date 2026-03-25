"""
Excel导出工具 - Coze插件简化版
只需要一个data参数即可
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io
import base64
from typing import List, Dict, Any
from datetime import datetime


def export_excel(data: List[Dict[str, Any]], filename: str = "export", sheet_name: str = "Sheet1", headers: Dict[str, str] = None) -> str:
    """
    导出数据到Excel并返回Base64字符串
    """
    if not data:
        raise ValueError("数据不能为空")

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    fields = list(data[0].keys())

    if headers:
        display_headers = [headers.get(field, field) for field in fields]
    else:
        display_headers = fields

    for col_num, header in enumerate(display_headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_num, row_data in enumerate(data, 2):
        for col_num, field in enumerate(fields, 1):
            value = row_data.get(field, '')
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            cell.font = Font(name='微软雅黑', size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    for col_num, header in enumerate(display_headers, 1):
        max_length = len(str(header)) + 2
        for row in worksheet.iter_rows(min_row=2, max_row=min(11, worksheet.max_row)):
            cell = row[col_num - 1]
            if cell.value:
                max_length = max(max_length, len(str(cell.value)) + 2)
        worksheet.column_dimensions[get_column_letter(col_num)].width = min(max(max_length, 10), 30)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return base64.b64encode(buffer.getvalue()).decode('utf-8')


async def handler(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Coze插件入口函数
    """
    try:
        data = args.get('data', [])

        if not data:
            return {
                'success': False,
                'message': '数据不能为空'
            }

        filename = args.get('filename') or f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        sheet_name = args.get('sheet_name') or 'Sheet1'
        headers = args.get('headers') or {}

        base64_content = export_excel(
            data=data,
            filename=filename,
            sheet_name=sheet_name,
            headers=headers if headers else None
        )

        return {
            'success': True,
            'message': 'Excel文件生成成功',
            'base64': base64_content,
            'filename': f'{filename}.xlsx',
            'record_count': len(data)
        }

    except Exception as e:
        return {
            'success': False,
            'message': f'导出失败：{str(e)}'
        }
