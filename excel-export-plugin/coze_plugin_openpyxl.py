"""
Excel导出工具 - Coze插件优化版
基于openpyxl，针对大数据量优化
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io
import base64
import json


def handler(args):
    """Coze插件入口函数"""
    try:
        # 获取参数
        raw_data = ''
        filename = 'export'
        headers_raw = '{}'
        sheet_name = 'Sheet1'
        
        # args.input.xxx 方式
        if hasattr(args, 'input'):
            inp = args.input
            raw_data = getattr(inp, 'data', '') or raw_data
            filename = getattr(inp, 'filename', '') or filename
            headers_raw = getattr(inp, 'headers', '{}') or headers_raw
            sheet_name = getattr(inp, 'sheet_name', 'Sheet1') or sheet_name
        
        # 直接属性
        if not raw_data:
            raw_data = getattr(args, 'data', '') or ''
            filename = getattr(args, 'filename', '') or 'export'
            headers_raw = getattr(args, 'headers', '{}') or '{}'
            sheet_name = getattr(args, 'sheet_name', 'Sheet1') or 'Sheet1'
        
        # 字典
        if not raw_data and isinstance(args, dict):
            raw_data = args.get('data', '')
            filename = args.get('filename', 'export')
            headers_raw = args.get('headers', '{}')
            sheet_name = args.get('sheet_name', 'Sheet1')
        
        if not raw_data:
            return {'success': False, 'message': '数据不能为空'}
        
        # 解析数据
        if isinstance(raw_data, str):
            data = json.loads(raw_data)
        else:
            data = raw_data
        
        if not isinstance(data, list) or len(data) == 0:
            return {'success': False, 'message': '数据格式错误'}
        
        # 解析headers
        if isinstance(headers_raw, str):
            headers = json.loads(headers_raw) if headers_raw else {}
        else:
            headers = headers_raw if headers_raw else {}
        
        # 字段
        fields = list(data[0].keys())
        
        if headers:
            display_headers = [headers.get(f, f) for f in fields]
        else:
            display_headers = fields
        
        record_count = len(data)
        
        # 根据数据量选择导出方式
        # 小数据量（<=10条）：生成Excel下载链接
        # 大数据量（>10条）：返回CSV文本，让用户复制保存
        if record_count <= 10:
            return _export_excel_url(data, fields, display_headers, filename, sheet_name)
        else:
            return _export_csv_text(data, fields, display_headers, filename, record_count)
        
    except Exception as e:
        return {'success': False, 'message': f'错误: {str(e)}'}


def _export_excel_url(data, fields, display_headers, filename, sheet_name):
    """小数据量：生成Excel下载链接"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(sheet_name)[:31]
    
    # 表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # 写入表头
    for col, header in enumerate(display_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col, field in enumerate(fields, 1):
            ws.cell(row=row_idx, column=col, value=row_data.get(field, ''))
    
    # 保存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    download_url = f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}"
    
    return {
        'success': True,
        'message': 'Excel文件生成成功，点击链接下载',
        'filename': filename + '.xlsx',
        'record_count': len(data),
        'download_url': download_url,
        'csv_content': '',
        'export_type': 'url'
    }


def _export_csv_text(data, fields, display_headers, filename, record_count):
    """大数据量：返回CSV文本，让用户复制保存"""
    lines = [','.join(display_headers)]
    
    for row_data in data:
        values = []
        for f in fields:
            v = str(row_data.get(f, '')) if row_data.get(f) is not None else ''
            if ',' in v or '"' in v or '\n' in v:
                v = '"' + v.replace('"', '""') + '"'
            values.append(v)
        lines.append(','.join(values))
    
    csv_content = '\n'.join(lines)
    
    return {
        'success': True,
        'message': f'数据量较大({record_count}条)，请复制下方CSV内容保存',
        'filename': filename + '.csv',
        'record_count': record_count,
        'download_url': '',
        'csv_content': csv_content,
        'export_type': 'text'
    }
