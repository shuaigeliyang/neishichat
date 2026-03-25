"""
Excel导出工具 - Coze插件精简版
解决循环引用问题
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io
import base64
import json
from datetime import datetime


async def handler(args: dict) -> dict:
    """
    Coze插件入口函数
    """
    try:
        raw_data = args.get('data', '')
        
        if not raw_data:
            return {
                'success': False,
                'message': '数据不能为空'
            }
        
        if isinstance(raw_data, str):
            data = json.loads(raw_data)
        else:
            data = raw_data
        
        if not isinstance(data, list) or len(data) == 0:
            return {
                'success': False,
                'message': '数据格式错误，需要数组格式'
            }
        
        filename = args.get('filename') or f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        sheet_name = args.get('sheet_name') or 'Sheet1'
        
        headers_raw = args.get('headers', '{}')
        if isinstance(headers_raw, str):
            headers = json.loads(headers_raw) if headers_raw else {}
        else:
            headers = headers_raw if headers_raw else {}
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name[:31]
        
        fields = list(data[0].keys())
        
        if headers:
            display_headers = [headers.get(field, field) for field in fields]
        else:
            display_headers = fields
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col_num, header in enumerate(display_headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=str(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
        
        data_alignment = Alignment(horizontal='left', vertical='center')
        data_font = Font()
        even_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        for row_num, row_data in enumerate(data, 2):
            for col_num, field in enumerate(fields, 1):
                value = row_data.get(field, '')
                cell = worksheet.cell(row=row_num, column=col_num, value=str(value) if value is not None else '')
                cell.font = data_font
                cell.alignment = data_alignment
                if row_num % 2 == 0:
                    cell.fill = even_row_fill
        
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        base64_content = base64.b64encode(buffer.getvalue()).decode('utf-8')
        
        workbook.close()
        
        return {
            'success': True,
            'message': 'Excel文件生成成功',
            'base64': base64_content,
            'filename': f'{filename}.xlsx',
            'record_count': len(data)
        }
        
    except json.JSONDecodeError as e:
        return {
            'success': False,
            'message': f'JSON格式错误：{str(e)}'
        }
    except Exception as e:
        return {
            'success': False,
            'message': f'导出失败：{str(e)}'
        }
