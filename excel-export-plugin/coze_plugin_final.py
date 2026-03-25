"""
Excel导出工具 - Coze插件最简版
只需要一个data参数（JSON字符串格式）
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io
import base64
import json
from datetime import datetime


async def handler(args: dict) -> dict:
    """
    Coze插件入口函数
    
    参数:
        data: JSON字符串，例如 '[{"学号":"001","姓名":"张三"}]'
              或者直接传入数组对象
    
    返回:
        success: 是否成功
        message: 结果消息
        base64: Excel文件的Base64编码
        filename: 文件名
    """
    try:
        raw_data = args.get('data', '')
        
        if not raw_data:
            return {
                'success': False,
                'message': '数据不能为空'
            }
        
        if isinstance(raw_data, str):
            data = json.loads(raw_data)
        else:
            data = raw_data
        
        if not isinstance(data, list) or len(data) == 0:
            return {
                'success': False,
                'message': '数据格式错误，需要数组格式'
            }
        
        filename = args.get('filename') or f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        sheet_name = args.get('sheet_name') or 'Sheet1'
        headers = args.get('headers') or {}
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        
        fields = list(data[0].keys())
        
        if headers:
            display_headers = [headers.get(field, field) for field in fields]
        else:
            display_headers = fields
        
        for col_num, header in enumerate(display_headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_num, row_data in enumerate(data, 2):
            for col_num, field in enumerate(fields, 1):
                value = row_data.get(field, '')
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                cell.font = Font(name='微软雅黑', size=11)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        for col_num, header in enumerate(display_headers, 1):
            max_length = len(str(header)) + 2
            for row in worksheet.iter_rows(min_row=2, max_row=min(11, worksheet.max_row)):
                cell = row[col_num - 1]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)) + 2)
            worksheet.column_dimensions[get_column_letter(col_num)].width = min(max(max_length, 10), 30)
        
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        base64_content = base64.b64encode(buffer.getvalue()).decode('utf-8')
        
        return {
            'success': True,
            'message': 'Excel文件生成成功',
            'base64': base64_content,
            'filename': f'{filename}.xlsx',
            'record_count': len(data)
        }
        
    except json.JSONDecodeError:
        return {
            'success': False,
            'message': 'JSON格式错误，请检查数据格式'
        }
    except Exception as e:
        return {
            'success': False,
            'message': f'导出失败：{str(e)}'
        }
