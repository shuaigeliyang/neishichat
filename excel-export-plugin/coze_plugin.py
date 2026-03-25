"""
Excel导出工具 - Coze插件版
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io
import base64
from typing import List, Dict, Any
from datetime import datetime


class ExcelExporter:
    """Excel导出工具类"""

    def __init__(self, filename: str = None):
        self.filename = filename or f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.workbook = None
        self.worksheet = None

    def export_to_excel(self,
                       data: List[Dict[str, Any]],
                       sheet_name: str = "Sheet1",
                       headers: Dict[str, str] = None,
                       auto_column_width: bool = True) -> str:
        """
        导出数据到Excel并返回Base64字符串

        Args:
            data: 要导出的数据列表
            sheet_name: 工作表名称
            headers: 列头映射
            auto_column_width: 是否自动调整列宽

        Returns:
            Base64编码的Excel文件内容
        """
        if not data:
            raise ValueError("数据不能为空")

        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name

        fields = list(data[0].keys())

        if headers:
            display_headers = [headers.get(field, field) for field in fields]
        else:
            display_headers = fields

        self._write_header(display_headers)
        self._write_data(data, fields)

        if auto_column_width:
            self._auto_adjust_column_width(fields, display_headers)

        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)

        return base64.b64encode(buffer.getvalue()).decode('utf-8')

    def _write_header(self, headers: List[str]):
        """写入表头"""
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _write_data(self, data: List[Dict[str, Any]], fields: List[str]):
        """写入数据行"""
        for row_num, row_data in enumerate(data, 2):
            for col_num, field in enumerate(fields, 1):
                value = row_data.get(field, '')
                cell = self.worksheet.cell(row=row_num, column=col_num, value=value)
                cell.font = Font(name='微软雅黑', size=11)
                cell.alignment = Alignment(horizontal='left', vertical='center')

                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    def _auto_adjust_column_width(self, fields: List[str], display_headers: List[str]):
        """自动调整列宽"""
        for col_num, (field, header) in enumerate(zip(fields, display_headers), 1):
            max_length = len(str(header)) + 2

            for row in self.worksheet.iter_rows(min_row=2, max_row=min(11, self.worksheet.max_row)):
                cell = row[col_num - 1]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)) + 2)

            adjusted_width = min(max(max_length, 10), 30)
            self.worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width


async def handler(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Coze插件入口函数

    Args:
        args: 插件输入参数
            - data: 要导出的数据列表
            - filename: 文件名（可选）
            - sheet_name: 工作表名称（可选）
            - headers: 表头映射（可选）

    Returns:
        导出结果
            - success: 是否成功
            - message: 结果消息
            - base64: Excel文件的Base64编码
            - filename: 文件名
    """
    try:
        data = args.get('data', [])
        filename = args.get('filename', f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
        sheet_name = args.get('sheet_name', 'Sheet1')
        headers = args.get('headers', {})

        if not data:
            return {
                'success': False,
                'message': '数据不能为空'
            }

        exporter = ExcelExporter(filename)
        base64_content = exporter.export_to_excel(
            data=data,
            sheet_name=sheet_name,
            headers=headers
        )

        return {
            'success': True,
            'message': 'Excel文件生成成功',
            'base64': base64_content,
            'filename': f'{filename}.xlsx',
            'record_count': len(data)
        }

    except Exception as e:
        return {
            'success': False,
            'message': f'导出失败：{str(e)}'
        }
