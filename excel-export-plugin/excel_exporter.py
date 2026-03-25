#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel导出工具 - Coze插件专用版
功能：提供高性能、灵活的Excel导出功能
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import pandas as pd
from typing import List, Dict, Any, Optional
import os
import io


class ExcelExporter:
    """
    Excel导出工具类

    功能特点：
    - 支持自定义样式（表头、行高、列宽）
    - 支持大数据量导出（性能优化）
    - 支持多工作表
    - 自动调整列宽
    - 支持返回Buffer（适合Web API）
    """

    def __init__(self, filename: str = None):
        """
        初始化Excel导出器

        Args:
            filename: 导出的文件名（不含扩展名）
        """
        self.filename = filename or f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.workbook = None
        self.worksheet = None

    def export_to_excel(self,
                       data: List[Dict[str, Any]],
                       sheet_name: str = "Sheet1",
                       headers: Dict[str, str] = None,
                       auto_column_width: bool = True,
                       style_config: Dict[str, Any] = None,
                       return_buffer: bool = False) -> Any:
        """
        导出数据到Excel文件

        Args:
            data: 要导出的数据列表（字典格式）
            sheet_name: 工作表名称
            headers: 列头映射（数据库字段 -> 显示名称）
            auto_column_width: 是否自动调整列宽
            style_config: 样式配置
            return_buffer: 是否返回Buffer而不是保存文件

        Returns:
            如果return_buffer=True，返回Buffer对象；否则返回文件路径
        """
        if not data:
            raise ValueError("数据不能为空")

        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name

        fields = list(data[0].keys())

        if headers:
            display_headers = [headers.get(field, field) for field in fields]
        else:
            display_headers = fields

        self._write_header(display_headers, style_config)
        self._write_data(data, fields, style_config)

        if auto_column_width:
            self._auto_adjust_column_width(fields, display_headers)

        if return_buffer:
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            return buffer
        else:
            return self._save_file()

    def export_with_pandas(self,
                          data: List[Dict[str, Any]],
                          sheet_name: str = "Sheet1",
                          headers: Dict[str, str] = None,
                          include_index: bool = False,
                          return_buffer: bool = False) -> Any:
        """
        使用Pandas导出（适合大数据量）

        Args:
            data: 要导出的数据
            sheet_name: 工作表名称
            headers: 列头映射
            include_index: 是否包含索引列
            return_buffer: 是否返回Buffer

        Returns:
            Buffer或文件路径
        """
        if not data:
            raise ValueError("数据不能为空")

        df = pd.DataFrame(data)

        if headers:
            df = df.rename(columns=headers)

        if return_buffer:
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=include_index)
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                self._apply_pandas_style(worksheet, df.shape[1])
            buffer.seek(0)
            return buffer
        else:
            filepath = f"{self.filename}.xlsx"
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=include_index)
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                self._apply_pandas_style(worksheet, df.shape[1])
            return filepath

    def export_multiple_sheets(self,
                              data_dict: Dict[str, List[Dict[str, Any]]],
                              headers_dict: Dict[str, Dict[str, str]] = None,
                              return_buffer: bool = False) -> Any:
        """
        导出多个工作表

        Args:
            data_dict: {工作表名: 数据列表}
            headers_dict: {工作表名: 列头映射}
            return_buffer: 是否返回Buffer

        Returns:
            Buffer或文件路径
        """
        self.workbook = openpyxl.Workbook()
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)

        for sheet_name, data in data_dict.items():
            if not data:
                continue

            worksheet = self.workbook.create_sheet(title=sheet_name)
            fields = list(data[0].keys())
            headers = headers_dict.get(sheet_name, {}) if headers_dict else {}
            display_headers = [headers.get(field, field) for field in fields]

            self._write_header_to_sheet(worksheet, display_headers)
            self._write_data_to_sheet(worksheet, data, fields)
            self._auto_adjust_column_width_for_sheet(worksheet, fields, display_headers)

        if return_buffer:
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            return buffer
        else:
            return self._save_file()

    def _write_header(self, headers: List[str], style_config: Dict[str, Any] = None):
        """写入表头"""
        style_config = style_config or self._get_default_header_style()

        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(
                name=style_config.get('font_name', '微软雅黑'),
                size=style_config.get('font_size', 12),
                bold=style_config.get('bold', True),
                color=style_config.get('font_color', 'FFFFFF')
            )
            cell.fill = PatternFill(
                start_color=style_config.get('bg_color', '4472C4'),
                end_color=style_config.get('bg_color', '4472C4'),
                fill_type='solid'
            )
            cell.alignment = Alignment(
                horizontal=style_config.get('align', 'center'),
                vertical=style_config.get('valign', 'center'),
                wrap_text=True
            )

    def _write_data(self, data: List[Dict[str, Any]], fields: List[str], style_config: Dict[str, Any] = None):
        """写入数据行"""
        row_num = 2

        for row_data in data:
            for col_num, field in enumerate(fields, 1):
                value = row_data.get(field, '')
                cell = self.worksheet.cell(row=row_num, column=col_num, value=value)
                cell.font = Font(name='微软雅黑', size=11)
                cell.alignment = Alignment(horizontal='left', vertical='center')

                if row_num % 2 == 0:
                    cell.fill = PatternFill(
                        start_color='F2F2F2',
                        end_color='F2F2F2',
                        fill_type='solid'
                    )

            row_num += 1

    def _write_header_to_sheet(self, worksheet, headers: List[str]):
        """向指定工作表写入表头"""
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _write_data_to_sheet(self, worksheet, data: List[Dict[str, Any]], fields: List[str]):
        """向指定工作表写入数据"""
        for row_num, row_data in enumerate(data, 2):
            for col_num, field in enumerate(fields, 1):
                value = row_data.get(field, '')
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                cell.font = Font(name='微软雅黑', size=11)
                cell.alignment = Alignment(horizontal='left', vertical='center')

                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    def _auto_adjust_column_width(self, fields: List[str], display_headers: List[str]):
        """自动调整列宽"""
        for col_num, (field, header) in enumerate(zip(fields, display_headers), 1):
            max_length = len(header) + 2

            for row in self.worksheet.iter_rows(min_row=2, max_row=min(11, self.worksheet.max_row)):
                cell = row[col_num - 1]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)) + 2)

            adjusted_width = min(max(max_length, 10), 30)
            self.worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    def _auto_adjust_column_width_for_sheet(self, worksheet, fields: List[str], display_headers: List[str]):
        """为指定工作表自动调整列宽"""
        for col_num, (field, header) in enumerate(zip(fields, display_headers), 1):
            max_length = len(header) + 2

            for row in worksheet.iter_rows(min_row=2, max_row=min(11, worksheet.max_row)):
                cell = row[col_num - 1]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)) + 2)

            adjusted_width = min(max(max_length, 10), 30)
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    def _apply_pandas_style(self, worksheet, num_columns: int):
        """为Pandas导出的工作表应用样式"""
        for col_num in range(1, num_columns + 2):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _save_file(self) -> str:
        """保存文件"""
        if not self.filename.endswith('.xlsx'):
            filepath = f"{self.filename}.xlsx"
        else:
            filepath = self.filename

        export_dir = 'exports'
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)

        full_path = os.path.join(export_dir, filepath)
        self.workbook.save(full_path)

        return full_path

    def _get_default_header_style(self) -> Dict[str, Any]:
        """获取默认表头样式配置"""
        return {
            'font_name': '微软雅黑',
            'font_size': 12,
            'bold': True,
            'font_color': 'FFFFFF',
            'bg_color': '4472C4',
            'align': 'center',
            'valign': 'center'
        }


def quick_export(data: List[Dict[str, Any]],
                filename: str = None,
                sheet_name: str = "Sheet1",
                return_buffer: bool = False) -> Any:
    """
    快速导出Excel - 最简单的方式

    Args:
        data: 要导出的数据
        filename: 文件名
        sheet_name: 工作表名
        return_buffer: 是否返回Buffer

    Returns:
        Buffer或文件路径
    """
    exporter = ExcelExporter(filename)
    return exporter.export_to_excel(data, sheet_name=sheet_name, return_buffer=return_buffer)


if __name__ == "__main__":
    test_data = [
        {"学号": "2021001", "姓名": "张三", "性别": "男", "学院": "计算机科学与技术学院", "专业": "软件工程"},
        {"学号": "2021002", "姓名": "李四", "性别": "女", "学院": "计算机科学与技术学院", "专业": "人工智能"},
        {"学号": "2021003", "姓名": "王五", "性别": "男", "学院": "软件工程学院", "专业": "物联网工程"},
    ]

    print("测试基础导出...")
    filepath = quick_export(test_data, "test_export", sheet_name="测试数据")
    print(f"导出成功: {filepath}")
