"""
Excel导出工具 - Coze插件简化版
直接生成Excel文件，返回下载链接
依赖：openpyxl
"""

import io
import base64
import json

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except:
    HAS_OPENPYXL = False


def handler(args):
    """Coze插件入口函数"""
    try:
        # 获取参数
        raw_data = ''
        filename = 'export'
        headers_raw = '{}'
        sheet_name = 'Sheet1'
        
        # args.input.xxx 方式
        if hasattr(args, 'input'):
            inp = args.input
            raw_data = getattr(inp, 'data', '') or raw_data
            filename = getattr(inp, 'filename', '') or filename
            headers_raw = getattr(inp, 'headers', '{}') or headers_raw
            sheet_name = getattr(inp, 'sheet_name', 'Sheet1') or sheet_name
        
        # 直接属性
        if not raw_data:
            raw_data = getattr(args, 'data', '') or ''
            filename = getattr(args, 'filename', '') or 'export'
            headers_raw = getattr(args, 'headers', '{}') or '{}'
            sheet_name = getattr(args, 'sheet_name', 'Sheet1') or 'Sheet1'
        
        # 字典
        if not raw_data and isinstance(args, dict):
            raw_data = args.get('data', '')
            filename = args.get('filename', 'export')
            headers_raw = args.get('headers', '{}')
            sheet_name = args.get('sheet_name', 'Sheet1')
        
        if not raw_data:
            return {'success': False, 'message': '数据不能为空'}
        
        # 解析数据
        if isinstance(raw_data, str):
            data = json.loads(raw_data)
        else:
            data = raw_data
        
        if not isinstance(data, list) or len(data) == 0:
            return {'success': False, 'message': '数据格式错误'}
        
        # 解析headers
        if isinstance(headers_raw, str):
            headers = json.loads(headers_raw) if headers_raw else {}
        else:
            headers = headers_raw if headers_raw else {}
        
        # 字段
        fields = list(data[0].keys())
        
        if headers:
            display_headers = [headers.get(f, f) for f in fields]
        else:
            display_headers = fields
        
        # 使用openpyxl生成Excel
        if HAS_OPENPYXL:
            return _export_excel(data, fields, display_headers, filename, sheet_name)
        else:
            return _export_csv(data, fields, display_headers, filename)
        
    except Exception as e:
        return {'success': False, 'message': f'错误: {str(e)}'}


def _export_excel(data, fields, display_headers, filename, sheet_name):
    """使用openpyxl导出Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(sheet_name)[:31]
    
    # 表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # 写入表头
    for col, header in enumerate(display_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col, field in enumerate(fields, 1):
            ws.cell(row=row_idx, column=col, value=row_data.get(field, ''))
    
    # 保存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    download_url = f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}"
    
    return {
        'success': True,
        'message': 'Excel文件生成成功',
        'filename': filename + '.xlsx',
        'record_count': len(data),
        'download_url': download_url
    }


def _export_csv(data, fields, display_headers, filename):
    """导出CSV（无openpyxl时使用）"""
    lines = [','.join(display_headers)]
    
    for row_data in data:
        values = []
        for f in fields:
            v = str(row_data.get(f, '')) if row_data.get(f) is not None else ''
            if ',' in v or '"' in v or '\n' in v:
                v = '"' + v.replace('"', '""') + '"'
            values.append(v)
        lines.append(','.join(values))
    
    content = '\n'.join(lines)
    b64 = base64.b64encode(content.encode('utf-8-sig')).decode('utf-8')
    download_url = f"data:text/csv;base64,{b64}"
    
    return {
        'success': True,
        'message': 'CSV文件生成成功',
        'filename': filename + '.csv',
        'record_count': len(data),
        'download_url': download_url
    }
